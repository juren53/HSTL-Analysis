"""
Reads 'Assets by Row Numbers-*.xlsx', looks up each Collection # in the
Collections tab, and writes the matching Collection Name into column B of
the Assets tab.  Saves the result as a new file so the original is untouched.
"""

# --- Imports ---
import sys                 # command-line arguments (sys.argv) and sys.exit()
from pathlib import Path   # filesystem path handling
import openpyxl            # read/write .xlsx workbooks

# --- Default file paths (overridable via command-line arguments) ---
INPUT_FILE = "Assets by Row Numbers-2026-06-24-09-12-13.xlsx"
OUTPUT_FILE = "Assets by Row Numbers-filled.xlsx"

# --- Sheet (tab) names inside the workbook ---
ASSETS_SHEET = "Assets by Row Numbers"
COLLECTIONS_SHEET = "Collections"

# Collections tab layout (1-based column numbers)
COL_ID   = 1   # Column A — Collection ID  (e.g. "HST-PCBP")
COL_NAME = 2   # Column B — Collection Name (e.g. "Brown, Peter C. Papers")

# Assets tab layout
ASSETS_COL_ID   = 1   # Column A — Collection #
ASSETS_COL_NAME = 2   # Column B — where the name should go


def build_lookup(ws) -> dict[str, str]:
    """Return {collection_id: collection_name} from the Collections sheet."""
    lookup = {}
    # Skip the header row (row 1); read every collection ID/name pair below it.
    for row in ws.iter_rows(min_row=2, values_only=True):
        coll_id   = row[COL_ID - 1]
        coll_name = row[COL_NAME - 1]
        if coll_id:
            # Store a blank name (rather than skipping) so a known ID with no
            # name is later distinguishable from a lookup miss only in intent,
            # not behavior — both are treated as "no match" by fill_names().
            lookup[str(coll_id).strip()] = str(coll_name).strip() if coll_name else ""
    return lookup


def fill_names(input_path: Path, output_path: Path) -> None:
    # --- Load the workbook and grab both sheets we need ---
    wb = openpyxl.load_workbook(input_path)

    ws_collections = wb[COLLECTIONS_SHEET]
    ws_assets      = wb[ASSETS_SHEET]

    # --- Build the Collection ID -> Collection Name lookup table ---
    lookup = build_lookup(ws_collections)
    print(f"Loaded {len(lookup):,} collection entries from '{COLLECTIONS_SHEET}' tab.")

    filled = 0
    missing = []

    # --- Walk every asset row and fill in column B ---
    for row_num, row in enumerate(
        ws_assets.iter_rows(min_row=2), start=2
    ):
        coll_id_cell = row[ASSETS_COL_ID - 1]
        name_cell    = row[ASSETS_COL_NAME - 1]

        raw = coll_id_cell.value
        if raw is None:
            continue  # no Collection # on this row — leave it alone

        coll_id = str(raw).strip()
        name    = lookup.get(coll_id)

        if name:
            name_cell.value = name
            filled += 1
        else:
            # Covers both "ID not found" and "ID found but name is blank".
            name_cell.value = "#N/A"
            missing.append((row_num, coll_id))

    # --- Save to a new file; the input workbook is never modified ---
    wb.save(output_path)

    # --- Report results ---
    print(f"Filled {filled:,} rows.")
    if missing:
        print(f"\nNo match found for {len(missing)} Collection ID(s):")
        for row_num, coll_id in missing[:20]:
            print(f"  Row {row_num}: '{coll_id}'")
        if len(missing) > 20:
            print(f"  ... and {len(missing) - 20} more.")
    print(f"\nSaved to: {output_path}")


if __name__ == "__main__":
    # Accept optional input/output paths as command-line arguments, falling
    # back to the defaults above if they weren't supplied.
    input_path  = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(INPUT_FILE)
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(OUTPUT_FILE)

    if not input_path.exists():
        sys.exit(f"Error: '{input_path}' not found.")

    fill_names(input_path, output_path)
